# -*- coding: utf-8 -*-
"""
Kareem T-01 — EDECS Contract Processor  (local Windows app — backend engine)

For every bundle PDF (uploaded in the app or placed in Inbox):
  1. Reads project code / contract no / annex from the file name.
  2. OCRs the pages (bundled Tesseract), detects the Purchase-Order report and the
     contract copy that follows it.  Detection order: visual-repeat -> FRM footer ->
     contract-number token -> manual [s-e] range.
  3. Saves ONE contract copy as a correctly-named PDF in Output.
  4. Builds a fully-mapped 26-column SR Log row and writes it to a SEPARATE file
     ("SR Log - NEW ROWS.xlsx") — the shared master is read read-only for de-dup only.
  5. Builds an editable .eml draft (X-Unsent:1) or a real Outlook draft.

This version adds (per the build spec):
  * Bundled OCR via ocr_engine (system Tesseract is fallback only).
  * Selectable OCR language per run.
  * Per-page OCR confidence + <70% -> Manual Review routing.
  * Optional manual Agreement Date & SR Number (override OCR; never block on missing).
  * Missing-field highlighting flags + SR Log notes.
  * Structured audit/history log (audit_log.jsonl).
The original detection & enrichment logic is preserved.
"""
import os, re, json, glob, getpass, datetime, traceback

import ocr_engine as OCR

HERE = os.environ.get("KT01_ROOT") or os.path.dirname(os.path.abspath(__file__))

def load_config():
    p = os.path.join(HERE, "config.json")
    cfg = json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}
    if not cfg.get("sync_dir"):
        for base in [os.environ.get("OneDrive", ""), os.environ.get("OneDriveCommercial", ""),
                     os.path.join(os.path.expanduser("~"), "OneDrive")]:
            cand = os.path.join(base, "Procurement-Sync") if base else ""
            if cand and os.path.isdir(cand):
                cfg["sync_dir"] = cand; break
    cfg.setdefault("sync_dir", HERE)                       # default: app folder (local, offline)
    cfg.setdefault("ocr_mode", "bundled")
    cfg.setdefault("ocr_default_language", "ara+eng")
    cfg.setdefault("ocr_dpi", 300)
    cfg.setdefault("ocr_confidence_test", True)
    cfg.setdefault("ocr_force_bundled", False)             # if True, never fall back to system
    cfg.setdefault("name_separator", "_")
    cfg.setdefault("make_outlook_draft", True)
    cfg.setdefault("output_dir", "")                       # optional explicit output folder
    # robust: if SR Log isn't in sync_dir, look beside the app / its parent
    if not os.path.exists(os.path.join(cfg["sync_dir"], "SR Log 2026.xlsx")):
        for cand in (HERE, os.path.dirname(HERE)):
            if os.path.exists(os.path.join(cand, "SR Log 2026.xlsx")):
                cfg["sync_dir"] = cand; break
    return cfg

CFG    = load_config()
SYNC   = CFG["sync_dir"]
INBOX  = os.path.join(SYNC, "Inbox")
OUTPUT = CFG.get("output_dir") or os.path.join(SYNC, "Output")
SRLOG  = os.path.join(SYNC, "SR Log 2026.xlsx")
AUDIT  = os.path.join(OUTPUT, "audit_log.jsonl")
MRQ    = os.path.join(OUTPUT, "manual_review.json")   # persistent manual-review queue

REFS   = json.load(open(os.path.join(HERE, "refs.json"), encoding="utf-8"))
RECIPS = json.load(open(os.path.join(HERE, "recipients.json"), encoding="utf-8"))
KB, LABELS, AR_ENG = REFS["kb"], REFS["label"], REFS["ar_eng"]

MON_ABBR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
CONF_THRESHOLD = 70.0     # below this -> Manual Review (spec section 6)

WORK_MAP = [
    (r"equipment|equ\b", ("Equ. Rent", "Logistic - Equ Rent", "معدات")),
    (r"consult",         ("Consultant", "Consultant agreement", "خدمات استشارية")),
    (r"transport|waste", ("Logistic", "Logistic - Inland transportation", "نقل")),
    (r"survey",          ("Survey", "Subcontractor agreement", "أعمال مساحة")),
    (r"ready ?mix",      ("Ready Mix", "Subcontractor agreement", "خرسانة جاهزة")),
    (r"steel|formwork|metal", ("Steel", "Subcontractor agreement", "أعمال حدادة ونجارة")),
    (r"asphalt|road",    ("Asphalt", "Subcontractor agreement", "أعمال أسفلت وطرق")),
    (r"paint",           ("Painting", "Subcontractor agreement", "أعمال دهانات")),
    (r"plaster|masonry|gypsum", ("Finishing", "Subcontractor agreement", "أعمال تشطيبات")),
    (r"marble|floor|interlock", ("Finishing", "Subcontractor agreement", "أعمال أرضيات")),
    (r"pile|soil|earth|concrete", ("Civil", "Subcontractor agreement", "أعمال خرسانة وأساسات")),
    (r"dolomite|sand|aggregate|rock", ("Materials", "Subcontractor agreement", "توريد مواد")),
]
def map_work(w):
    if not w: return ("", "Subcontractor agreement", "")
    for pat, val in WORK_MAP:
        if re.search(pat, w.lower()): return val
    return (w, "Subcontractor agreement", w)

def ncode(c):
    """Normalise project code: strip leading zeros from base, zero-pad sub to 2, keep letters."""
    c = c.replace("_", "-"); ps = c.split("-")
    b = ps[0].lstrip("0") or "0"
    sub = [p.zfill(2) if p.isdigit() else p for p in ps[1:]]
    return "-".join([b] + sub) if sub else b

# Accepted filename examples shown to the user when parsing fails (no AI; deterministic).
PARSE_EXAMPLES = (
    "107-59659.pdf | 98-60033.pdf | "
    "72-49835 Annex(1).pdf | 78-48033 Annex (2).pdf | "
    "099-02-26700 Annex (1).pdf | CON-01-57697.pdf | "
    "099-02-26700 [3-4].pdf | 099-02-26700 Annex(1) [3-4].pdf"
)

def annex_label(n):
    """Normalise annex number to the canonical display form 'Annex (N)'."""
    return "Annex (%d)" % n if n else ""

def parse_bundle_name(fn):
    """
    Deterministic, AI-free bundle-name parser. Tolerant of spacing/case.

    Layout:  <projectCode>-<contractNo>[ Annex(N) | Annex (N)] [ [s-e] ]
      * projectCode = digits, optional '-NN' sub-project, or letter codes (CON-01, N-Store)
      * contractNo  = the LAST run of 3-6 digits that belongs to the code
      * Annex / addendum number optional, any case, with or without a space before '('
      * manual page range [s-e] optional
    Returns dict(code, no, annex, manual) or None if it cannot be parsed.
    """
    if not fn:
        return None
    name = os.path.splitext(str(fn))[0]          # ignore extension
    name = re.sub(r"\s+", " ", name).strip()      # trim / collapse spaces

    # 1) manual page range [s-e]  (also tolerates [s - e])
    manual = None
    mr = re.search(r"\[\s*(\d+)\s*-\s*(\d+)\s*\]", name)
    if mr:
        manual = (int(mr.group(1)), int(mr.group(2)))
    core = re.sub(r"\[\s*\d+\s*-\s*\d+\s*\]", " ", name)   # strip range token

    # 2) annex / addendum number  (Annex(1), Annex (1), ANNEX1, addendum 2, ...)
    annex = None
    ma = re.search(r"\b(?:annex|addendum)\b\s*\(?\s*(\d+)\s*\)?", core, re.I)
    if ma:
        annex = int(ma.group(1))
    core = re.sub(r"\b(?:annex|addendum)\b\s*\(?\s*\d*\s*\)?", " ", core, flags=re.I)
    core = re.sub(r"\s+", " ", core).strip(" -_")

    # 3) split <code>-<contractNo>: contractNo is the LAST 3-6 digit group.
    #    Everything before the final '-<digits>' is the project code.
    m = re.match(r"^(?P<code>.+?)[-_ ]+(?P<no>\d{3,6})$", core)
    if not m:
        # also accept a bare 'code no' with a space separator
        m = re.match(r"^(?P<code>.+?)\s+(?P<no>\d{3,6})$", core)
    if not m:
        return None
    code_raw = m.group("code").strip(" -_")
    no = m.group("no")
    if not code_raw:
        return None
    return dict(code=ncode(code_raw), no=no, annex=annex, manual=manual)

_fitz = None
def get_fitz():
    global _fitz
    if _fitz is None:
        import fitz; _fitz = fitz
    return _fitz

def _prep_ocr():
    OCR.configure_pytesseract(CFG)

def ocr_pages(pdf_path, lang=None, want_conf=False):
    """OCR the first <=8 pages (PO report is always near the front). Optionally collect conf."""
    import pytesseract
    from PIL import Image
    _prep_ocr()
    lang = lang or CFG["ocr_default_language"]
    cfgstr = OCR.tess_config(CFG, extra="--psm 6")
    doc = get_fitz().open(pdf_path)
    zoom = CFG["ocr_dpi"] / 72.0
    mat = get_fitz().Matrix(zoom, zoom)
    out = []; confs = []
    for i in range(min(len(doc), 8)):
        pix = doc[i].get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        # PO report parsing relies on English layout; keep eng for those tokens but honor lang otherwise
        out.append((i + 1, pytesseract.image_to_string(img, lang="eng", config=cfgstr)))
        if want_conf:
            confs.append(OCR.page_confidence(img, CFG, lang))
    doc.close()
    if want_conf:
        avg = round(sum(confs) / len(confs), 1) if confs else 0.0
        return out, avg
    return out

def parse_po(text):
    d = {}
    m = re.search(r"ED[EC]{1,2}G?-?0*([0-9]{3,})", text); d["po_no"] = m.group(1) if m else None
    m = re.search(r"(\d{4}-\d{2}-\d{2})", text); d["po_date"] = m.group(1) if m else None
    m = re.search(r"\b(SUB|SER|CU)-?\s?0*([0-9]{2,})", text)
    d["vendor_account"] = f"{m.group(1)}-{m.group(2).zfill(5)}" if m else None
    amts = [(float(a.replace(",", "")), s) for s, a in re.findall(r"(-?)\s*([0-9][0-9,]*\.\d{2})\s*EGP", text)]
    pos = [a for a, s in amts if s == "" and a > 0]; negs = [a for a, s in amts if s == "-"]
    if pos:
        gross = max(pos); net = gross - sum(abs(x) for x in negs)
        exp = [a for a in pos if abs(a - net) < 1 and a != gross]
        d["net"] = round(exp[0] if exp else net, 2)
    else:
        d["net"] = None
    m = re.search(r"Report generated:\s*(\d{2}/\d{2}/\d{4})", text); d["report_date"] = m.group(1) if m else None
    m = re.search(r"APPROVED BY[^\n]*\n\s*((?:Dr\.?\s*|Eng\.?\s*)?[A-Za-z]+\s+[A-Za-z]+)", text)
    d["approved_by"] = m.group(1).strip() if m else None
    m = re.search(r"ISSUED BY[^\n]*\n(.+)", text)
    d["issued_raw"] = m.group(1) if m else ""
    d["issued_by"] = " ".join(m.group(1).split()[-2:]) if m else None
    return d

# ---- SR Number extraction (spec section 8) -------------------------------
def extract_sr_number(pages):
    """Try to find an SR number token in the OCR'd pages, e.g. SR-2026-00125 / SR 125."""
    text = "\n".join(t for _, t in pages)
    for pat in (r"SR[-\s]?No\.?\s*[:.]?\s*(SR[-\s]?\d{2,4}[-\s]?\d{2,6})",
                r"\b(SR[-\s]?\d{4}[-\s]?\d{2,6})\b",
                r"SR[-\s]?No\.?\s*[:.]?\s*([0-9]{2,8})"):
        m = re.search(pat, text, re.I)
        if m:
            return re.sub(r"\s+", "-", m.group(1).strip())
    return ""

def page_vectors(pdf_path):
    import numpy as np
    from PIL import Image
    doc = get_fitz().open(pdf_path)
    mat = get_fitz().Matrix(70 / 72.0, 70 / 72.0)
    V = []
    for i in range(len(doc)):
        pix = doc[i].get_pixmap(matrix=mat)
        im = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("L").resize((36, 50))
        a = np.asarray(im, dtype=float); a = (a - a.mean()) / (a.std() + 1e-6)
        V.append(a.flatten())
    doc.close()
    return V

def best_repeat(V, thr=0.6):
    """First contract copy = a contiguous page block that repeats later (the 2nd copy)."""
    import numpy as np
    n = len(V)
    if n < 2: return None
    Lv = len(V[0]); best = None
    for s in range(n):
        for L in range(1, (n - s) // 2 + 1):
            avg = sum(float(np.dot(V[s + k], V[s + L + k]) / Lv) for k in range(L)) / L
            if avg > thr:
                score = L * avg
                if best is None or score > best[0]:
                    best = (score, s, L, avg)
    if not best: return None
    _, s, L, avg = best
    copies = 2; k = 2
    while s + (k + 1) * L <= n and \
          sum(float(np.dot(V[s + m], V[s + k * L + m]) / Lv) for m in range(L)) / L > thr:
        copies += 1; k += 1
    return dict(start=s, length=L, copies=copies, sim=avg)

def frm_flags(pdf_path):
    """True per page if the contract form footer (FRM-...) is on its bottom ~14% strip."""
    import pytesseract
    from PIL import Image
    _prep_ocr()
    cfgstr = OCR.tess_config(CFG, extra="--psm 6")
    doc = get_fitz().open(pdf_path)
    mat = get_fitz().Matrix(150 / 72.0, 150 / 72.0)
    flags = []
    for i in range(len(doc)):
        pix = doc[i].get_pixmap(matrix=mat)
        im = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        w, h = im.size
        strip = im.crop((0, int(h * 0.86), w, h))
        t = pytesseract.image_to_string(strip, lang="eng", config=cfgstr)
        flags.append(bool(re.search(r"FRM", t)))
    doc.close()
    return flags

def longest_frm_run(flags):
    n = len(flags); runs = []; i = 0
    while i < n:
        if flags[i]:
            j = i
            while j < n and (flags[j] or (j + 1 < n and flags[j + 1])):
                j += 1
            runs.append((i, j - 1)); i = j
        else:
            i += 1
    if not runs: return None
    return max(runs, key=lambda r: r[1] - r[0])

def token_run(pages, code, no):
    """Short annexes: the contract page shows 'no/code' (e.g. 26700/099-02)."""
    base = code.split("-")[0]
    tok = re.compile(rf"{no}\s*/\s*0*{base}|0*{base}\s*/\s*{no}")
    hits = [i for i, (_, t) in enumerate(pages) if tok.search(t)]
    if not hits:
        return None
    s = e = hits[0]
    for i in hits[1:]:
        if i == e + 1: e = i
        else: break
    return (s, e)

def contract_date(pdf_path, page_idx, lang=None):
    """Read the contract date from the contract's first page (Arabic+English OCR)."""
    import pytesseract
    from PIL import Image
    _prep_ocr()
    lang = lang or CFG["ocr_default_language"] or "ara+eng"
    cfgstr = OCR.tess_config(CFG, extra="--psm 6")
    try:
        doc = get_fitz().open(pdf_path)
        pix = doc[page_idx].get_pixmap(matrix=get_fitz().Matrix(150 / 72.0, 150 / 72.0))
        im = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()
        t = pytesseract.image_to_string(im, lang=lang, config=cfgstr)
    except Exception:
        return ("", "")
    m = re.search(r"(\d{4})\s*[/\-]\s*(\d{1,2})\s*[/\-]\s*(\d{1,2})", t)
    if m and 1 <= int(m.group(2)) <= 12:
        return (f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}", f"{MON_ABBR[int(m.group(2))]} {m.group(1)}")
    m = re.search(r"(\d{1,2})\s*[/\-]\s*(\d{1,2})\s*[/\-]\s*(\d{4})", t)
    if m and 1 <= int(m.group(2)) <= 12:
        return (f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}", f"{MON_ABBR[int(m.group(2))]} {m.group(3)}")
    return ("", "")

def extract_contract_fields(pdf_path, copy_pages, lang=None):
    """
    Dynamically OCR the FIRST page(s) of the selected contract range and pull contract
    fields. Returns {field: {"value": str, "conf": float}}. NOTHING is hardcoded per
    document — only generic Arabic/English label patterns are used. Low confidence or
    no-match leaves value "" so the caller can route the field to Manual Review.
    """
    import pytesseract
    from PIL import Image
    if not copy_pages:
        return {}
    _prep_ocr()
    lang = lang or CFG.get("ocr_default_language") or "ara+eng"
    cfgstr = OCR.tess_config(CFG, extra="--psm 6")
    doc = get_fitz().open(pdf_path)
    text = ""; pconf = []
    for idx in copy_pages[:2]:                      # first 1-2 pages carry the header info
        if idx >= len(doc):
            continue
        pix = doc[idx].get_pixmap(matrix=get_fitz().Matrix(200 / 72.0, 200 / 72.0))
        im = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        text += "\n" + pytesseract.image_to_string(im, lang=lang, config=cfgstr)
        pconf.append(OCR.page_confidence(im, CFG, lang))
    doc.close()
    base_conf = round(sum(pconf) / len(pconf), 1) if pconf else 0.0

    def grab(patterns):
        for pat in patterns:
            m = re.search(pat, text, re.I)
            if m:
                v = re.sub(r"\s+", " ", m.group(1)).strip(" :.-،")
                if v:
                    return v
        return ""

    fields = {}
    # supplier / second party (generic Arabic labels for "the second party / contractor")
    fields["company"] = grab([
        r"(?:الطرف\s*الثاني|المقاول|المورد|الشركة)\s*[:\-/]?\s*([^\n]{3,60})",
        r"(?:second\s*party|contractor|supplier|vendor)\s*[:\-/]?\s*([^\n]{3,60})",
    ])
    # scope / subject of the contract
    fields["scope"] = grab([
        r"(?:موضوع\s*(?:العقد|الاتفاق)|نطاق\s*العمل|الأعمال)\s*[:\-/]?\s*([^\n]{3,80})",
        r"(?:scope(?:\s*of\s*work)?|subject)\s*[:\-/]?\s*([^\n]{3,80})",
    ])
    # project name
    fields["project"] = grab([
        r"(?:اسم\s*المشروع|المشروع|مشروع)\s*[:\-/]?\s*([^\n]{3,60})",
        r"(?:project(?:\s*name)?)\s*[:\-/]?\s*([^\n]{3,60})",
    ])
    # PO / contract number tokens (generic)
    fields["po_no"] = grab([
        r"(ED[EC]{1,2}G?-?0*\d{3,})",
        r"(?:PO|أمر\s*شراء|رقم\s*العقد)\s*(?:No\.?|رقم)?\s*[:\-/]?\s*([A-Z0-9\-]{4,})",
    ])
    # amount / contract value (largest money-looking token)
    amts = re.findall(r"([0-9][0-9,]*\.\d{2})", text)
    fields["amount"] = max(amts, key=lambda a: float(a.replace(",", ""))) if amts else ""

    # attach a confidence to each (base OCR conf if we got a value, else 0)
    return {k: {"value": v, "conf": (base_conf if v else 0.0)} for k, v in fields.items()}, base_conf


def page_of_n_flags(pdf_path, lang="eng"):
    """
    OCR the bottom strip of every page and read a 'Page X of N' footer if present.
    Returns a list (per page) of (x, n) tuples or None. Tolerant of OCR noise and of
    a few Arabic/spacing variants (e.g. 'Page 1 of 7', '1 of 7', '1/7', 'صفحة 1 من 7').
    No contract content is hardcoded; only the page/total pattern is read.
    """
    import pytesseract
    from PIL import Image
    _prep_ocr()
    cfgstr = OCR.tess_config(CFG, extra="--psm 6")
    doc = get_fitz().open(pdf_path)
    mat = get_fitz().Matrix(180 / 72.0, 180 / 72.0)
    out = []
    pats = [
        r"page\s*(\d+)\s*of\s*(\d+)",
        r"\b(\d+)\s*of\s*(\d+)\b",
        r"\b(\d+)\s*/\s*(\d+)\b",
        r"(?:صفحة|صفحه)\s*(\d+)\s*(?:من|/)\s*(\d+)",
    ]
    for i in range(len(doc)):
        pix = doc[i].get_pixmap(matrix=mat)
        im = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        w, h = im.size
        strip = im.crop((0, int(h * 0.85), w, h))
        t = " ".join(pytesseract.image_to_string(strip, lang=lang, config=cfgstr).split())
        hit = None
        for pat in pats:
            m = re.search(pat, t, re.I)
            if m:
                x, n = int(m.group(1)), int(m.group(2))
                if 1 <= x <= n <= 60:    # sanity bounds; not a fixed contract length
                    hit = (x, n); break
        out.append(hit)
    doc.close()
    return out

def sequences_from_pageof(flags):
    """
    Given per-page (x, n) footers, return complete contract sequences as (start, end)
    0-based inclusive blocks where x runs 1..N over consecutive pages with the same N.
    """
    seqs = []; n = len(flags); i = 0
    while i < n:
        f = flags[i]
        if f and f[0] == 1:                      # a block that begins at 'Page 1 of N'
            total = f[1]; j = i; expect = 1
            while j < n and flags[j] and flags[j][1] == total and flags[j][0] == expect:
                expect += 1; j += 1
            got = j - i
            # accept if we saw the whole 1..N (allow the last page or two to be OCR-missed)
            if got >= max(1, total - 1):
                end = i + total - 1
                if end < n:
                    seqs.append((i, end))
                    i = end + 1; continue
        i += 1
    return seqs

def valid_start_page(flags, start):
    """
    A contract block must not start at 'Page 2 of N' when the previous page is
    'Page 1 of N' of the same sequence. Returns the corrected start (shift back to
    the real Page 1 if found), else the original start.
    """
    if start < 0 or start >= len(flags):
        return start
    here = flags[start]
    if here and here[0] >= 2 and start > 0:
        prev = flags[start - 1]
        if prev and prev[1] == here[1] and prev[0] == here[0] - 1:
            # walk back to Page 1 of N
            s = start
            while s > 0 and flags[s - 1] and flags[s - 1][1] == here[1] \
                  and flags[s - 1][0] == flags[s][0] - 1:
                s -= 1
            return s
    return start

def analyse(pdf_path, pages, code, no, lang=None):
    """
    Decide which pages form ONE contract copy. Generic (no per-file hardcoding).
    Signal priority:
      1. 'Page 1 of N .. Page N of N' footer sequence(s) -> take the FIRST complete one.
      2. Visual-repeat of a contiguous block (two identical copies) with an adaptive
         similarity threshold, then validate the start is a real first page.
      3. FRM footer run (single-copy multi-page).
      4. Contract-number token run (short annex).
    Returns dict(copy_pages, copies, po_text, confident, reason).
    'confident' False or copy_pages None -> caller routes to Manual Review.
    """
    lang = lang or CFG.get("ocr_default_language") or "ara+eng"
    po_text = "\n".join(t for _, t in pages if re.search(r"purch.{0,4}order", t, re.I))

    def _validate_start(start):
        """Lazily OCR footers only to confirm we didn't start at 'Page 2 of N'."""
        try:
            pof = page_of_n_flags(pdf_path, lang="eng")
        except Exception:
            return start, []
        return valid_start_page(pof, start), pof

    # (1) FAST visual repeat first (no OCR): two identical copies, adaptive threshold.
    V = page_vectors(pdf_path)
    rep = None
    for thr in (0.92, 0.88, 0.85, 0.82):
        rep = best_repeat(V, thr=thr)
        if rep:
            break
    if rep:
        s, L = rep["start"], rep["length"]
        s, _ = _validate_start(s)                 # shift back to a real Page 1 if needed
        return dict(copy_pages=list(range(s, s + L)), copies=rep["copies"],
                    po_text=po_text, confident=True,
                    reason="visual-repeat sim=%.3f" % rep["sim"])

    # (2) Page-1..N footer sequence (single copy or no clear repeat).
    try:
        pof = page_of_n_flags(pdf_path, lang="eng")
    except Exception:
        pof = []
    seqs = sequences_from_pageof(pof) if pof else []
    if seqs:
        s, e = seqs[0]
        s = valid_start_page(pof, s)
        return dict(copy_pages=list(range(s, e + 1)), copies=max(1, len(seqs)),
                    po_text=po_text, confident=True, reason="page-of-N sequence")

    # (3) FRM footer run (single copy, multi-page).
    blk = longest_frm_run(frm_flags(pdf_path))
    if blk and blk[1] > blk[0]:
        s = blk[0]
        if pof:
            s = valid_start_page(pof, s)
        return dict(copy_pages=list(range(s, blk[1] + 1)), copies=1,
                    po_text=po_text, confident=True, reason="FRM footer run")

    # (4) contract-number token run (short annex).
    tr = token_run(pages, code, no)
    if tr:
        return dict(copy_pages=list(range(tr[0], tr[1] + 1)), copies=1,
                    po_text=po_text, confident=True, reason="contract-number token")

    return dict(copy_pages=None, copies=0, po_text=po_text, confident=False,
                reason="No reliable contract page range found")

_SR = None
def sr_refs():
    global _SR
    if _SR is not None: return _SR
    import openpyxl
    try:
        wb = openpyxl.load_workbook(SRLOG, data_only=True)
    except Exception:
        _SR = dict(vend={}, scope={}, div={}, cat={}, names=set()); return _SR
    vend = {}
    if "Vendors" in wb.sheetnames:
        vs = wb["Vendors"]
        for r in range(2, vs.max_row + 1):
            a = vs.cell(r, 1).value
            if a: vend[str(a).strip().replace("_", "-")] = vs.cell(r, 2).value
    scope = {}; div = {}; cat = {}
    ws = wb["SR Log"]
    for r in range(2, ws.max_row + 1):
        a = str(ws.cell(r, 11).value or "").replace("_", "-")
        if not a: continue
        if a not in scope and ws.cell(r, 9).value: scope[a] = ws.cell(r, 9).value
        if a not in div and ws.cell(r, 13).value: div[a] = ws.cell(r, 13).value
        if a not in cat and ws.cell(r, 10).value: cat[a] = ws.cell(r, 10).value
    names = set()
    if "ERP User" in wb.sheetnames:
        eu = wb["ERP User"]
        for r in range(2, eu.max_row + 1):
            v = eu.cell(r, 2).value
            if v and isinstance(v, str) and len(v.split()) >= 2: names.add(v.strip())
    for d in RECIPS.values():
        for fld in ("to", "cc"):
            for nm in re.findall(r"([A-Za-z][A-Za-z]+(?:\s+[A-Za-z]+){1,3})\s*<", d.get(fld, "")):
                names.add(nm.strip())
    wb.close()
    _SR = dict(vend=vend, scope=scope, div=div, cat=cat, names=names)
    return _SR

def match_person(text, names):
    t = (text or "").lower(); best = ""; key = (-1, -1)
    for nm in names:
        i = t.rfind(nm.lower())
        if i >= 0 and (i, len(nm)) > key: key = (i, len(nm)); best = nm
    return best

def norm_ar(s):
    s = re.sub(r"[ًٌٍَُِّْـ‏‎]", "", s or "")
    for a, b in [("أ","ا"),("إ","ا"),("آ","ا"),("ة","ه"),("ى","ي")]:
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip()

def ar_to_eng(arabic):
    if not arabic: return ""
    na = norm_ar(arabic)
    for k, v in AR_ENG.items():
        if norm_ar(k) == na: return v
    a2 = " ".join(na.split()[:2])
    for k, v in AR_ENG.items():
        if " ".join(norm_ar(k).split()[:2]) == a2: return v
    return ""

def enrich(meta, po):
    code, no = meta["code"], meta["no"]
    ref = KB.get(f"{code}|{no}", {}); sr = sr_refs(); acct = po.get("vendor_account") or ""
    label = ref.get("label") or LABELS.get(code) or code
    arabic_excel = sr["vend"].get(acct) or ref.get("arabic") or ""
    arabic_file = ref.get("arabic") or arabic_excel
    scope = (ref.get("work") or sr["scope"].get(acct) or "").strip()
    veng = ref.get("veng") or ar_to_eng(arabic_file) or arabic_file
    cat = sr["cat"].get(acct) or map_work(scope)[1]
    div = sr["div"].get(acct) or map_work(scope)[2]
    person = match_person(po.get("issued_raw", ""), sr["names"]) or (po.get("issued_by") or "")
    return dict(code=code, no=no, annex=meta["annex"], label=label, veng=veng, work=scope,
                arabic_excel=arabic_excel, arabic_file=arabic_file, scope=scope, cat=cat, div=div, person=person)

def build_name(meta, po, en, my):
    head = "Agreement" + (f" Annex ({en['annex']})" if en["annex"] else "")
    mid = f"PJ{en['code']}_{en['no']} {en['label']}-{en['veng']}-{en['work']}-{my}".rstrip("-")
    sep = CFG["name_separator"]
    return (f"{head}-{mid}{sep}{en['arabic_file']}").rstrip(sep).strip()

def _safe(name):
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()

def write_contract_pdf(src, copy_pages, out_name):
    doc = get_fitz().open(src); new = get_fitz().open()
    for p in copy_pages:
        new.insert_pdf(doc, from_page=p, to_page=p)
    dest = os.path.join(OUTPUT, _safe(out_name) + ".pdf")
    new.save(dest); new.close(); doc.close()
    return dest

SR_COLS = 26
def build_row(meta, po, copies, en, date_iso, sr_number, note_extra=""):
    proj = LABELS.get(en["code"], en["label"])
    typ = f"Addendum {en['annex']}" if en["annex"] else "Contract"
    rd = po.get("report_date") or ""
    note = (po.get("approved_by") or "")
    if note_extra:
        note = (note + " | " + note_extra).strip(" |")
    row = [""] * SR_COLS
    row[0]=en["code"]; row[1]=proj; row[3]=sr_number or ""; row[4]=date_iso; row[7]=en["person"]
    row[8]=en["scope"]; row[9]=en["cat"]; row[10]=po.get("vendor_account") or ""; row[11]=en["arabic_excel"]
    row[12]=en["div"]; row[13]=en["no"]; row[14]=typ; row[15]=date_iso
    row[16]=po.get("net") if po.get("net") is not None else ""
    row[19]=rd; row[20]=rd; row[21]="Open"; row[22]=note; row[23]=copies
    return row

NEW_ROWS_NAME = "SR Log - NEW ROWS.xlsx"
def _master_keys():
    """Read the (possibly server-locked) master read-only for de-dup; safe if open elsewhere."""
    import openpyxl
    seen = set()
    try:
        m = openpyxl.load_workbook(SRLOG, read_only=True); ws = m["SR Log"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[0]: seen.add((str(r[0] or ""), str(r[13] or ""), str(r[14] or "")))
        m.close()
    except Exception:
        pass
    return seen

def append_sr_row(row, highlight_cols=None):
    """Write new rows to a SEPARATE file in Output (never locks/touches the shared master).
    highlight_cols: list of 0-based column indexes to shade yellow (missing fields)."""
    import openpyxl
    from openpyxl.styles import PatternFill
    out = os.path.join(OUTPUT, NEW_ROWS_NAME)
    seen = _master_keys()
    if os.path.exists(out):
        nb = openpyxl.load_workbook(out); ns = nb["New Rows"]
        for r in ns.iter_rows(min_row=2, values_only=True):
            if r and r[0]: seen.add((str(r[0] or ""), str(r[13] or ""), str(r[14] or "")))
    else:
        nb = openpyxl.Workbook(); ns = nb.active; ns.title = "New Rows"
        hdr = None
        try:
            m = openpyxl.load_workbook(SRLOG, read_only=True); ws = m["SR Log"]
            hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]; m.close()
        except Exception:
            pass
        ns.append(hdr or DEFAULT_HEADERS)
    if (str(row[0]), str(row[13]), str(row[14])) in seen:
        nb.close(); return False
    ns.append(row)
    if highlight_cols:
        fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
        r = ns.max_row
        for c in highlight_cols:
            ns.cell(row=r, column=c + 1).fill = fill
    nb.save(out); nb.close(); return True

DEFAULT_HEADERS = ["Project Code","Project Name","Request Type","No.","Create Date",
    "Cost Control Sign Date","Procurement Sign Date","Person","Scope","Category",
    "Vendor account","Vendor","Division","PO / Contract No.","Type","Date","Amount",
    "Procurement Sign Date2","Cost Control Sign Date2","Chairman Sign Date","Distribution Date",
    "Status","Note","Archive Copies No.","Archive Vendor","Archive Date"]

def recipients_for(code, work):
    r = RECIPS.get(code) or RECIPS.get(code.split("-")[0]) or {"to": "", "cc": ""}
    cc = [p.strip() for p in re.split(r";\s*", r.get("cc", "")) if p.strip()]
    if "equipment" not in (work or "").lower():
        cc = [p for p in cc if "logistics" not in p.lower()]
    return r.get("to", ""), "; ".join(cc)

def email_body(proj, code, veng, work, copies):
    cp = {1:"One",2:"Two",3:"Three",4:"Four"}.get(copies, str(copies))
    return ("Dear All,\r\n\r\nKindly find attached Subcontractor's Contract detailed as follows:-\r\n\r\n"
            f"Project          : {proj} (PJ {code}).\r\n"
            f"Contractor       : {veng}.\r\n"
            f"Contract Subject : {work}.\r\n"
            f"Copies           : {cp} Original(s).\r\n\r\n"
            "This is for your information and Record.\r\n\r\nBest Regards,\r\n")

def make_draft(name, to, cc, subject, body, pdf_path):
    name = _safe(name)
    if CFG["make_outlook_draft"]:
        try:
            import win32com.client
            ol = win32com.client.Dispatch("Outlook.Application")
            mail = ol.CreateItem(0)
            mail.To = to; mail.CC = cc; mail.Subject = subject; mail.Body = body
            mail.Attachments.Add(os.path.abspath(pdf_path)); mail.Save()
            return "Outlook draft", ""
        except Exception:
            pass
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["X-Unsent"] = "1"; msg["To"] = to; msg["Cc"] = cc; msg["Subject"] = subject
    msg.set_content(body)
    with open(pdf_path, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename=name + ".pdf")
    eml = os.path.join(OUTPUT, name + ".eml")
    open(eml, "wb").write(bytes(msg))
    return ".eml file", eml

# ---- Audit log (spec section 17) -----------------------------------------
def audit(record):
    os.makedirs(OUTPUT, exist_ok=True)
    record = dict(record)
    record.setdefault("timestamp", datetime.datetime.now().isoformat(timespec="seconds"))
    record.setdefault("user", getpass.getuser())
    with open(AUDIT, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")

def read_audit():
    if not os.path.exists(AUDIT): return []
    out = []
    for line in open(AUDIT, encoding="utf-8"):
        line = line.strip()
        if line:
            try: out.append(json.loads(line))
            except Exception: pass
    return out


# ---- Persistent Manual Review queue (spec: ISSUE 1) -----------------------
def _mrq_load():
    if not os.path.exists(MRQ):
        return {}
    try:
        return json.load(open(MRQ, encoding="utf-8"))
    except Exception:
        return {}

def _mrq_save(data):
    os.makedirs(OUTPUT, exist_ok=True)
    json.dump(data, open(MRQ, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def mrq_upsert(record):
    """Create/update a manual-review record keyed by file id (filename stem)."""
    data = _mrq_load()
    fid = record["file"]
    prev = data.get(fid, {})
    prev.update(record)
    prev["updated"] = datetime.datetime.now().isoformat(timespec="seconds")
    data[fid] = prev
    _mrq_save(data)
    return prev

def mrq_get(fid):
    return _mrq_load().get(fid)

def mrq_list():
    return list(_mrq_load().values())

def mrq_resolve(fid):
    """Mark a record corrected (move out of the active queue)."""
    data = _mrq_load()
    if fid in data:
        data[fid]["status"] = "Corrected"
        data[fid]["updated"] = datetime.datetime.now().isoformat(timespec="seconds")
        _mrq_save(data)
    return data.get(fid)

# ---------------------------------------------------------------------------
# MAIN: process a single bundle. Returns a rich dict for the API/UI.
#   status: "done" | "warning" | "manual" | "failed"
# ---------------------------------------------------------------------------
def process_one(pdf_path, lang=None, manual_agreement_date="", manual_sr_number="",
                manual_range=None, manual=None):
    """
    Process one bundle. `manual` (dict) may carry user corrections from the Manual Review
    page and OVERRIDES OCR for: agreement_date, sr_number, po_no, vendor_account, company,
    scope, amount, project, range_start, range_end. Missing/weak fields never block; they
    create a persistent Manual Review record and continue (status 'warning').
    """
    manual = dict(manual or {})
    if manual_agreement_date and not manual.get("agreement_date"): manual["agreement_date"] = manual_agreement_date
    if manual_sr_number and not manual.get("sr_number"): manual["sr_number"] = manual_sr_number
    if manual.get("range_start") and manual.get("range_end") and not manual_range:
        manual_range = (int(manual["range_start"]), int(manual["range_end"]))

    fn = os.path.splitext(os.path.basename(pdf_path))[0]
    lang = lang or CFG["ocr_default_language"]
    result = {"file": fn, "status": "failed", "msg": "", "confidence": None,
              "missing": [], "warnings": [], "output_pdf": "", "eml": "",
              "sr_added": False, "review_reason": "",
              "extracted_fields": {}, "missing_fields": [], "low_confidence_fields": [],
              "manual_review_required": False, "manual_review_reason": "",
              "page_range": None,
              "parsed": {"original": os.path.basename(pdf_path), "code": "", "no": "",
                         "annex": None, "type": "", "status": "Needs Manual Review"}}

    meta = parse_bundle_name(fn)
    if not meta:
        result.update(status="manual", manual_review_required=True,
                      review_reason=("Filename not understood. Accepted examples: " + PARSE_EXAMPLES),
                      manual_review_reason=("Filename not understood. Accepted examples: " + PARSE_EXAMPLES),
                      msg="اسم الملف غير مفهوم. أمثلة مقبولة: " + PARSE_EXAMPLES)
        mrq_upsert({"file": fn, "original": os.path.basename(pdf_path), "status": "Needs Manual Review",
                    "reason": result["manual_review_reason"], "extracted": {},
                    "missing": ["filename"], "low_confidence": [],
                    "code": "", "no": "", "annex": None, "range_start": "", "range_end": ""})
        audit({"file": fn, "status": "manual", "manual_review_reason": result["manual_review_reason"]})
        return result

    result["parsed"].update(
        code=meta["code"], no=meta["no"], annex=meta["annex"],
        type=("Addendum %d" % meta["annex"] if meta["annex"] else "Contract"),
        status="Parsed")

    try:
        pages, conf = ocr_pages(pdf_path, lang=lang, want_conf=True)
    except Exception as e:
        result.update(status="failed", msg="OCR error: %s" % e)
        audit({"file": fn, "status": "failed", "msg": result["msg"], "ocr_language": lang})
        return result
    result["confidence"] = conf

    # contract detection (manual range overrides everything)
    rng = manual_range or meta.get("manual")
    if rng:
        s_, e_ = rng
        po_text = "\n".join(t for _, t in pages if re.search(r"purch.{0,4}order", t, re.I))
        info = {"copy_pages": list(range(s_ - 1, e_)), "copies": 1, "po_text": po_text,
                "confident": True, "reason": "manual range [%d-%d]" % (s_, e_)}
    else:
        info = analyse(pdf_path, pages, meta["code"], meta["no"], lang=lang)

    # If detection failed OR OCR confidence is below threshold AND no manual range given,
    # route to Manual Review (do not silently finalise).
    if (not info or not info.get("copy_pages") or not info.get("confident")):
        reason = (info.get("reason") if info else "") or "Contract page range not detected."
        result.update(status="manual", manual_review_required=True,
                      review_reason=reason + " Set start/end pages in Manual Review.",
                      manual_review_reason=reason + " Set start/end pages in Manual Review.",
                      msg="تعذّر تحديد صفحات العقد — حدّد الصفحات في المراجعة اليدوية")
        mrq_upsert({"file": fn, "original": os.path.basename(pdf_path), "status": "Needs Manual Review",
                    "reason": result["manual_review_reason"], "extracted": {},
                    "missing": ["page_range"], "low_confidence": [],
                    "code": meta["code"], "no": meta["no"], "annex": meta["annex"],
                    "range_start": "", "range_end": "", "ocr_confidence": conf})
        audit({"file": fn, "status": "manual", "ocr_confidence": conf, "ocr_language": lang,
               "manual_review_reason": result["manual_review_reason"]})
        return result

    copy_pages = info["copy_pages"]
    result["page_range"] = {"start": copy_pages[0] + 1, "end": copy_pages[-1] + 1,
                            "count": len(copy_pages), "reason": info.get("reason", "")}

    po = parse_po(info["po_text"]) if info["po_text"] else {}
    en = enrich(meta, po)

    # ---- Dynamic field extraction from the SELECTED contract range -------
    try:
        xf, range_conf = extract_contract_fields(pdf_path, copy_pages, lang=lang)
    except Exception:
        xf, range_conf = {}, 0.0

    # ---- Agreement Date (manual overrides OCR) --------------------------
    ci, cmy = contract_date(pdf_path, copy_pages[0], lang=lang)
    po_my = ""
    if po.get("po_date"):
        y, m, _ = po["po_date"].split("-"); po_my = f"{MON_ABBR[int(m)]} {y}"
    detected_date = ci
    agreement_date = (manual.get("agreement_date") or "").strip() or detected_date or po.get("po_date") \
                     or KB.get(f"{meta['code']}|{meta['no']}", {}).get("my", "")
    date_iso = agreement_date or ""
    my = ""
    if manual.get("agreement_date"):
        mm = re.search(r"(\d{1,2})\D+(\d{4})", manual["agreement_date"]) or \
             re.search(r"(\d{4})\D+(\d{1,2})", manual["agreement_date"])
        if mm:
            g = mm.groups()
            month = int(g[0]) if len(g[0]) <= 2 else int(g[1])
            year = g[1] if len(g[1]) == 4 else g[0]
            if 1 <= month <= 12: my = f"{MON_ABBR[month]} {year}"
    my = my or cmy or po_my or KB.get(f"{meta['code']}|{meta['no']}", {}).get("my", "")

    # ---- SR Number (manual overrides OCR) -------------------------------
    detected_sr = extract_sr_number(pages)
    sr_number = (manual.get("sr_number") or "").strip() or detected_sr

    # ---- resolve each business field: enrichment -> OCR extraction -> manual
    def pick(manual_key, enriched, ocr_key):
        mv = (manual.get(manual_key) or "").strip()
        if mv: return mv, "manual"
        if enriched: return enriched, "enriched"
        ov = (xf.get(ocr_key, {}) or {}).get("value", "")
        if ov: return ov, "ocr"
        return "", "none"

    vendor_account = (manual.get("vendor_account") or "").strip() or (po.get("vendor_account") or "")
    company, company_src = pick("company", en.get("arabic_excel") or en.get("veng"), "company")
    scope, scope_src     = pick("scope", en.get("scope"), "scope")
    amount_manual = (manual.get("amount") or "").strip()
    net = None
    if amount_manual:
        try: net = float(amount_manual.replace(",", ""))
        except Exception: net = None
    if net is None:
        net = po.get("net")
    if net is None and xf.get("amount", {}).get("value"):
        try: net = float(xf["amount"]["value"].replace(",", ""))
        except Exception: net = None
    po_no = (manual.get("po_no") or "").strip() or po.get("po_no") or (xf.get("po_no", {}) or {}).get("value", "")

    result["extracted_fields"] = {
        "company": company, "scope": scope, "vendor_account": vendor_account,
        "po_no": po_no, "amount": (("%.2f" % net) if net is not None else ""),
        "agreement_date": date_iso, "sr_number": sr_number,
        "project": (manual.get("project") or en.get("label") or (xf.get("project", {}) or {}).get("value", "")),
        "range_conf": range_conf,
    }

    # ---- missing / low-confidence flags ---------------------------------
    notes = []; highlight = []
    def flag_missing(key, label, cols):
        result["missing_fields"].append(key); result["warnings"].append(label + " missing / not detected")
        notes.append(label + " missing / not detected"); highlight.extend(cols)
    if not date_iso:       flag_missing("agreement_date", "Agreement Date", [4, 15])
    if not sr_number:      flag_missing("sr_number", "SR Number", [3])
    if not vendor_account: flag_missing("vendor_account", "Vendor account", [10])
    if net is None:        flag_missing("amount", "Amount", [16])
    if not company:        flag_missing("company", "Company / supplier", [11])
    if not scope:          flag_missing("scope", "Scope", [8])
    # weak OCR overall or per-range
    if conf and conf < CONF_THRESHOLD:
        result["low_confidence_fields"].append("ocr_overall")
    if range_conf and range_conf < CONF_THRESHOLD:
        result["low_confidence_fields"].append("contract_range_ocr")
    result["missing"] = result["missing_fields"]   # backward-compat for the Process table

    # ---- outputs (always produced; manual values already merged) ---------
    name = build_name(meta, po, en, my)
    # ensure SR Log uses resolved values even when they came from OCR/manual
    en2 = dict(en); en2["arabic_excel"] = company or en.get("arabic_excel"); en2["scope"] = scope or en.get("scope")
    po2 = dict(po); po2["vendor_account"] = vendor_account or po.get("vendor_account"); po2["net"] = net
    pdf_out = write_contract_pdf(pdf_path, copy_pages, name)
    added = append_sr_row(build_row(meta, po2, info["copies"], en2, date_iso, sr_number,
                                    note_extra=" | ".join(notes)),
                          highlight_cols=sorted(set(highlight)))
    proj = LABELS.get(meta["code"], en["label"])
    to, cc = recipients_for(meta["code"], scope or en["scope"])
    body = email_body(proj, meta["code"], company or en["veng"], scope or en["scope"], info["copies"])
    kind, eml = make_draft(name, to, cc, name, body, pdf_out)

    # ---- manual-review record if anything is missing/weak ---------------
    need_review = bool(result["missing_fields"] or result["low_confidence_fields"])
    result["manual_review_required"] = need_review
    if need_review:
        result["manual_review_reason"] = "Fields need confirmation: " + ", ".join(
            result["missing_fields"] + result["low_confidence_fields"])
        mrq_upsert({
            "file": fn, "original": os.path.basename(pdf_path),
            "status": "Needs Manual Review",
            "reason": result["manual_review_reason"],
            "code": meta["code"], "no": meta["no"], "annex": meta["annex"],
            "range_start": copy_pages[0] + 1, "range_end": copy_pages[-1] + 1,
            "ocr_confidence": conf, "range_conf": range_conf,
            "extracted": result["extracted_fields"],
            "missing": result["missing_fields"], "low_confidence": result["low_confidence_fields"],
            "output_pdf": os.path.basename(pdf_out), "eml": os.path.basename(eml) if eml else "",
        })
    else:
        # if it was previously queued and is now clean, mark corrected
        if mrq_get(fn):
            mrq_resolve(fn)

    status = "warning" if (result["warnings"] or need_review) else "done"
    result.update(status=status, output_pdf=os.path.basename(pdf_out), eml=eml,
                  sr_added=added, copies=info["copies"],
                  msg="عقد: %s | صفحات: %d-%d (%d) | نسخ: %d | SR Log: %s | إيميل: %s%s" % (
                      os.path.basename(pdf_out), copy_pages[0] + 1, copy_pages[-1] + 1, len(copy_pages),
                      info["copies"], "أُضيف" if added else "موجود", kind,
                      (" | مراجعة: " + "؛ ".join(result["warnings"])) if result["warnings"] else ""))

    audit({"file": fn, "output_pdf": result["output_pdf"], "project_code": meta["code"],
           "contract_no": meta["no"], "ocr_language": lang, "ocr_confidence": conf,
           "page_range": result["page_range"], "extracted_fields": result["extracted_fields"],
           "detected_agreement_date": detected_date, "manual_agreement_date": manual.get("agreement_date", ""),
           "detected_sr_number": detected_sr, "manual_sr_number": manual.get("sr_number", ""),
           "missing_fields": result["missing_fields"], "low_confidence_fields": result["low_confidence_fields"],
           "manual_review_required": need_review, "status": status,
           "email_draft": eml or kind, "sr_log_output": os.path.join(OUTPUT, NEW_ROWS_NAME)})
    for k in ("agreement_date", "sr_number", "vendor_account", "company", "scope", "amount", "po_no"):
        if manual.get(k):
            audit({"field_changed": k, "manual_value": manual[k], "file": fn})
    return result


def run(log=print, progress=None, on_result=None, lang=None):
    """Batch-process every PDF in Inbox (advanced / legacy folder mode)."""
    os.makedirs(OUTPUT, exist_ok=True)
    log(f"المجلد: {SYNC}")
    if not os.path.exists(SRLOG):
        log(f"تنبيه: لم يتم العثور على {SRLOG} — سأكمل وأكتب الصفوف في Output بدون إثراء.")
    files = sorted(glob.glob(os.path.join(INBOX, "*.pdf")))
    total = len(files)
    if not files:
        log("لا توجد ملفات في Inbox.")
        if progress: progress(0, 0)
        return
    done = warn = manual = failed = 0
    for i, f in enumerate(files, 1):
        try:
            r = process_one(f, lang=lang)
        except Exception:
            r = {"status": "failed", "file": os.path.basename(f),
                 "msg": "خطأ: " + traceback.format_exc().splitlines()[-1]}
        st = r["status"]
        if st == "done": done += 1; log("[OK] " + r["file"] + " — " + r["msg"])
        elif st == "warning": warn += 1; log("[!] " + r["file"] + " — " + r["msg"])
        elif st == "manual":
            manual += 1; log("[?] " + r["file"] + " — " + r.get("review_reason", r["msg"]))
            with open(os.path.join(OUTPUT, "_pending.txt"), "a", encoding="utf-8") as p:
                p.write(f"{r['file']}\t{r.get('review_reason', r['msg'])}\n")
        else:
            failed += 1; log("[X] " + r["file"] + " — " + r["msg"])
        if on_result: on_result(r["file"], st, r.get("msg", ""))
        if progress: progress(i, total)
    log(f"\nانتهى. تمت {done} | تنبيهات {warn} | مراجعة {manual} | فشل {failed}.")

if __name__ == "__main__":
    run()
